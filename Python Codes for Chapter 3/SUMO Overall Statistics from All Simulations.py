import os
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
def parse_expected_by_type(file_path):
    expected_ids = set()
    expected_by_type = defaultdict(set)
    tree = ET.parse(file_path)
    root = tree.getroot()
    for trip in root.findall(".//trip"):
        vid = trip.get("id")
        vtype = trip.get("type") or trip.get("vType")
        if vid and vtype:
            base_type = vtype.strip().lower().split('@')[0]
            expected_ids.add(vid.strip())
            expected_by_type[base_type].add(vid.strip())
    return expected_ids, expected_by_type
def parse_vehicle_ids_from_file(file_path):
    vehicle_ids = set()
    tree = ET.parse(file_path)
    root = tree.getroot()
    for veh in root.findall(".//vehicle"):
        vid = veh.get("id")
        if vid:
            vehicle_ids.add(vid.strip())
    return vehicle_ids
def parse_Trip-Info_extended_metrics(file_path):
    completed_ids = set()
    completed_by_type = defaultdict(set)
    metrics = {
        "time_loss": defaultdict(list),
        "travel_time": defaultdict(list),
        "waiting_time": defaultdict(list),
        "route_length": defaultdict(list),
        "delay": defaultdict(list),
        "speed": defaultdict(list)
    }
    tree = ET.parse(file_path)
    root = tree.getroot()
    for trip in root.findall(".//Trip-Info"):
        vid = trip.get("id")
        vtype = trip.get("type") or trip.get("vType")
        if not vtype or not vid:
            continue
        base_type = vtype.strip().lower().split('@')[0]
        try:
            duration = float(trip.get("duration", 0))
            waiting_time = float(trip.get("waitingTime", 0))
            route_length = float(trip.get("routeLength", 0)) / 1000
            time_loss = float(trip.get("timeLoss", 0))
            delay = (time_loss - waiting_time) / 60
            duration_hr = duration / 3600
            speed = (route_length / duration_hr) if duration_hr > 0 else 0
            completed_ids.add(vid.strip())
            completed_by_type[base_type].add(vid.strip())
            metrics["time_loss"][base_type].append(time_loss / 60)
            metrics["travel_time"][base_type].append(duration / 60)
            metrics["waiting_time"][base_type].append(waiting_time / 60)
            metrics["route_length"][base_type].append(route_length)
            metrics["delay"][base_type].append(delay)
            metrics["speed"][base_type].append(speed)
        except:
            continue
    return completed_ids, completed_by_type, metrics
def compute_avg(metric_dict):
    return {k: round(sum(v) / len(v), 2) if v else 0.0 for k, v in metric_dict.items()}

def compute_weighted_time_loss(time_loss_dict):
    total_weighted_sum = 0
    total_completed = 0
    for vt in time_loss_dict:
        count = len(time_loss_dict[vt])
        avg = sum(time_loss_dict[vt]) / count if count > 0 else 0
        total_weighted_sum += avg * count
        total_completed += count
    return round(total_weighted_sum / total_completed, 2) if total_completed > 0 else 0.0
def compute_total_kilometers(route_length_dict):
    return {vt: round(sum(route_length_dict[vt]), 2) for vt in route_length_dict}
def compute_weighted_avg_speed(route_lengths, durations):
    total_length = sum(sum(route_lengths[vt]) for vt in route_lengths)
    total_duration = sum(sum(durations[vt]) for vt in durations)
    total_duration_hr = total_duration / 60
    return round(total_length / total_duration_hr, 2) if total_duration_hr > 0 else 0.0
def main():
    group_names = [
        "dynamicinitial", "dynamicstep1", "dynamicstep2", "dynamicstep3", "dynamicstep4", "dynamicstep5", "dynamicstep6",
        "dynamicraininitial", "dynamicrainstep1", "dynamicrainstep2", "dynamicrainstep3", "dynamicrainstep4", "dynamicrainstep5", "dynamicrainstep6",
        "dynamicpoststep1", "dynamicpoststep2+dynamicsposttep2"
    ]

    num_simulations = 10
    odtrips_file = "odtrips_merged.xml"
    output_excel = "customized_summary_output.xlsx"
    if not os.path.exists(odtrips_file):
        sys.exit(f"ERROR: '{odtrips_file}' not found.")
    expected_ids, expected_by_type = parse_expected_by_type(odtrips_file)
    for group in group_names:
        print(f"\n🔄 Processing group: {group}")
        summary_rows = []
        for i in range(1, num_simulations + 1):
            label = f"Sim_{i}"
            parts = group.split("+")
            Trip-Info_files = [f"{p}Trip-Info_{i}.xml" for p in parts]
            Veh-Rou_files = [f"{p}Veh-Rou_{i}.xml" for p in parts]
            inserted_ids = set()
            completed_ids = set()
            all_metrics = {m: defaultdict(list) for m in ["time_loss", "travel_time", "waiting_time", "route_length", "delay", "speed"]}
            completed_by_type = defaultdict(set)
            for vf in Veh-Rou_files:
                if os.path.exists(vf):
                    inserted_ids.update(parse_vehicle_ids_from_file(vf))
            for tf in Trip-Info_files:
                if os.path.exists(tf):
                    completed_set, comp_types, metrics = parse_Trip-Info_extended_metrics(tf)
                    completed_ids.update(completed_set)
                    for vt in comp_types:
                        completed_by_type[vt].update(comp_types[vt])
                    for m in all_metrics:
                        for vt, values in metrics[m].items():
                            all_metrics[m][vt].extend(values)
            avg_metrics = {m: compute_avg(all_metrics[m]) for m in all_metrics}
            weighted_avg_time_loss = compute_weighted_time_loss(all_metrics["time_loss"])
            weighted_avg_speed = compute_weighted_avg_speed(all_metrics["route_length"], all_metrics["travel_time"])
            total_kilometers = compute_total_kilometers(all_metrics["route_length"])
            total_km_all = round(sum(total_kilometers.values()), 2)
            not_inserted = expected_ids - inserted_ids
            inserted_not_completed = inserted_ids - completed_ids
            completed = completed_ids
            row = {
                "Simulation": label,
                "Total Expected Vehicles": len(expected_ids),
                "Total Inserted Vehicles": len(inserted_ids),
                "Total Completed Vehicles": len(completed),
                "Total Missing Vehicles": len(not_inserted),
                "Total Inserted Not Completed": len(inserted_not_completed),
                "Weighted Avg TimeLoss (min)": weighted_avg_time_loss,
                "Overall Avg Speed (km/h)": weighted_avg_speed,
                "Total Kilometers Traveled (All)": total_km_all
            }
            for vt in ["passenger", "truck", "bus"]:
                row[f"Expected {vt.title()}"] = len(expected_by_type.get(vt, set()))
                row[f"Completed {vt.title()}"] = len(completed_by_type.get(vt, set()))
                row[f"Missing {vt.title()}"] = len(expected_by_type.get(vt, set()) - completed_by_type.get(vt, set()))
                row[f"Avg TimeLoss {vt.title()} (min)"] = avg_metrics["time_loss"].get(vt, 0.0)
                row[f"Avg Travel Time {vt.title()} (min)"] = avg_metrics["travel_time"].get(vt, 0.0)
                row[f"Avg Waiting Time {vt.title()} (min)"] = avg_metrics["waiting_time"].get(vt, 0.0)
                row[f"Avg Route Length {vt.title()} (km)"] = avg_metrics["route_length"].get(vt, 0.0)
                row[f"Avg Speed {vt.title()} (km/h)"] = avg_metrics["speed"].get(vt, 0.0)
                row[f"Avg Delay {vt.title()} (min)"] = avg_metrics["delay"].get(vt, 0.0)
                row[f"Total Kilometers {vt.title()}"] = total_kilometers.get(vt, 0.0)
            summary_rows.append(row)
        df = pd.DataFrame(summary_rows)
        mean_row = df.drop(columns=["Simulation"]).mean(numeric_only=True)
        mean_row["Simulation"] = "Mean"
        df = pd.concat([df, pd.DataFrame([mean_row])], ignore_index=True)
        numeric_cols = df.select_dtypes(include='number').columns
        mean_values = df[df["Simulation"] == "Mean"][numeric_cols].iloc[0]
        distances = df[df["Simulation"] != "Mean"][numeric_cols].sub(mean_values).abs().sum(axis=1)
        min_distance = distances.min()
        closest_rows = distances[distances == min_distance].index.tolist()
        sheet = group.split("+")[0]
        mode = "a" if os.path.exists(output_excel) else "w"
        writer_args = {"engine": "openpyxl", "mode": mode}
        if mode == "a":
            writer_args["if_sheet_exists"] = "replace"
        with pd.ExcelWriter(output_excel, **writer_args) as writer:
            df.to_excel(writer, sheet_name=f"{sheet}_Summary", index=False)
        wb = load_workbook(output_excel)
        ws = wb[f"{sheet}_Summary"]
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        for row in closest_rows:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row + 2, column=col).fill = green_fill
        wb.save(output_excel)
    print(f"\n✅ Final customized summary saved to '{output_excel}'")
if __name__ == "__main__":
    main()
